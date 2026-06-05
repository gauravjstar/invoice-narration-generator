"""Excel and CSV export service"""
import logging
from typing import List, Dict, Optional
from datetime import datetime
from pathlib import Path
import csv

logger = logging.getLogger(__name__)

class ExportService:
    """Export invoice data to Excel and CSV formats"""
    
    EXPORT_COLUMNS = [
        ('File Name', 'file_name', 30),
        ('Vendor Name', 'vendor_name', 25),
        ('GSTIN', 'vendor_gstin', 15),
        ('Invoice No', 'invoice_number', 15),
        ('Invoice Date', 'invoice_date', 12),
        ('Taxable Amount', 'taxable_amount', 15),
        ('CGST', 'cgst_amount', 12),
        ('SGST', 'sgst_amount', 12),
        ('IGST', 'igst_amount', 12),
        ('Total Amount', 'total_amount', 15),
        ('Transporter', 'transporter_name', 20),
        ('Vehicle No', 'vehicle_number', 15),
        ('LR No', 'lr_number', 12),
        ('Narration', 'narration', 80),
    ]
    
    @staticmethod
    def export_to_excel(invoices: List[Dict], output_path: Optional[str] = None) -> str:
        """Export invoices to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            if output_path is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f"invoices_export_{timestamp}.xlsx"
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoices"
            
            # Write headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_idx, (header_name, _, width) in enumerate(ExportService.EXPORT_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = width
            
            # Write data
            for row_idx, invoice in enumerate(invoices, 2):
                for col_idx, (_, field_name, _) in enumerate(ExportService.EXPORT_COLUMNS, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Get value
                    if field_name in invoice:
                        value = invoice[field_name]
                        
                        # Format amounts
                        if field_name.endswith('_amount'):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                        elif field_name in ['invoice_date']:
                            if isinstance(value, datetime):
                                value = value.strftime('%d-%m-%Y')
                            cell.alignment = Alignment(horizontal='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', wrap_text=True)
                        
                        cell.value = value
                    
                    cell.border = border
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Save file
            wb.save(output_path)
            logger.info(f"✅ Excel exported to {output_path}")
            return output_path
            
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return ExportService.export_to_csv(invoices, output_path.replace('.xlsx', '.csv'))
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise
    
    @staticmethod
    def export_to_csv(invoices: List[Dict], output_path: Optional[str] = None) -> str:
        """Export invoices to CSV format"""
        try:
            if output_path is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f"invoices_export_{timestamp}.csv"
            
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = [field_name for _, field_name, _ in ExportService.EXPORT_COLUMNS]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                # Write header
                writer.writerow({field: field for field in fieldnames})
                
                # Write data
                for invoice in invoices:
                    row = {}
                    for _, field_name, _ in ExportService.EXPORT_COLUMNS:
                        value = invoice.get(field_name, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%d-%m-%Y')
                        row[field_name] = value
                    writer.writerow(row)
            
            logger.info(f"✅ CSV exported to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}")
            raise
    
    @staticmethod
    def export_summary(invoices: List[Dict]) -> Dict:
        """Generate export summary with statistics"""
        try:
            total_invoices = len(invoices)
            total_amount = sum(inv.get('total_amount', 0) for inv in invoices)
            total_taxable = sum(inv.get('taxable_amount', 0) for inv in invoices)
            total_cgst = sum(inv.get('cgst_amount', 0) for inv in invoices)
            total_sgst = sum(inv.get('sgst_amount', 0) for inv in invoices)
            total_igst = sum(inv.get('igst_amount', 0) for inv in invoices)
            
            return {
                'total_invoices': total_invoices,
                'total_amount': total_amount,
                'total_taxable': total_taxable,
                'total_cgst': total_cgst,
                'total_sgst': total_sgst,
                'total_igst': total_igst,
                'total_tax': total_cgst + total_sgst + total_igst,
                'average_amount': total_amount / total_invoices if total_invoices > 0 else 0,
            }
        except Exception as e:
            logger.error(f"Summary generation failed: {str(e)}")
            return {}
